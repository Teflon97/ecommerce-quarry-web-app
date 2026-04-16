from django.shortcuts import render
from django.contrib.auth.decorators import login_required
from django.http import HttpResponse
from apps.orders.models import Order
import openpyxl
from datetime import datetime

@login_required
def report_list(request):
    return render(request, 'admin/portal/reports/list.html')

@login_required
def export_orders(request):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Orders Report"
    
    # Headers
    headers = ['Order #', 'Customer Name', 'Email', 'Phone', 'Total Amount', 'Status', 'Date']
    for col, header in enumerate(headers, 1):
        ws.cell(row=1, column=col, value=header)
    
    # Data
    orders = Order.objects.all()
    for row, order in enumerate(orders, 2):
        ws.cell(row=row, column=1, value=order.order_number)
        ws.cell(row=row, column=2, value=order.customer_name)
        ws.cell(row=row, column=3, value=order.customer_email)
        ws.cell(row=row, column=4, value=order.customer_phone)
        ws.cell(row=row, column=5, value=float(order.total_amount))
        ws.cell(row=row, column=6, value=order.get_status_display())
        ws.cell(row=row, column=7, value=order.created_at.strftime('%Y-%m-%d'))
    
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = f'attachment; filename="orders_{datetime.now().strftime("%Y%m%d")}.xlsx"'
    wb.save(response)
    return response